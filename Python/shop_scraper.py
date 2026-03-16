"""
Daraz Shop Scraper — Windows GUI Edition v2
Auto-installs dependencies, then launches a clean modern GUI.
Run: python daraz_scraper_gui.py
"""

import subprocess, sys, os

def _install(pkg):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

try:
    import aiohttp
except ImportError:
    print('Installing aiohttp...'); _install('aiohttp'); import aiohttp

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('Installing openpyxl...'); _install('openpyxl')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

import asyncio
import json
import random
import re
import threading
import tkinter as tk
from collections import defaultdict
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

# ══════════════════════════ SCRAPER CORE ══════════════════════════════

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.daraz.com.bd/",
}

HEADER_FILL  = PatternFill("solid", start_color="1F3864")
SUBHEAD_FILL = PatternFill("solid", start_color="2E75B6")
ALT_FILL     = PatternFill("solid", start_color="EBF3FB")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
ACCENT_FILL  = PatternFill("solid", start_color="D6E4F0")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name="Arial", bold=True, color="1F3864", size=14)
NORMAL_FONT  = Font(name="Arial", size=10)
LINK_FONT    = Font(name="Arial", size=10, color="0563C1", underline="single")
THIN_BORDER  = Border(
    left=Side(style="thin", color="BDD7EE"),   right=Side(style="thin", color="BDD7EE"),
    top=Side(style="thin",  color="BDD7EE"),   bottom=Side(style="thin", color="BDD7EE"),
)
XL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
XL_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
XL_RIGHT  = Alignment(horizontal="right",  vertical="center")


def build_ajax_url(shop_slug, page):
    is_first = "true" if page == 1 else "false"
    return (f"https://www.daraz.com.bd/{shop_slug}/"
            f"?ajax=true&from=wangpu&isFirstRequest={is_first}"
            f"&langFlag=en&page={page}&pageTypeId=2&q=All-Products&sort=order")


def parse_sold(s):
    if not s: return None
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def parse_products(data, shop_slug):
    try:
        items = data["mods"]["listItems"]
    except (KeyError, TypeError):
        return []
    products = []
    for item in items:
        if item.get("tItemType") != "nt_product": continue
        if not item.get("itemSoldCntShow"):        continue
        url = item.get("itemUrl", "")
        if url.startswith("//"): url = "https:" + url
        products.append({
            "shop":           shop_slug,
            "item_id":        item.get("itemId") or item.get("nid"),
            "name":           item.get("name", ""),
            "brand":          item.get("brandName", "") or "N/A",
            "seller":         item.get("sellerName", ""),
            "price":          int(item["price"]) if item.get("price") else None,
            "original_price": int(item["originalPrice"]) if item.get("originalPrice") else None,
            "discount":       item.get("discount", ""),
            "rating":         float(item["ratingScore"]) if item.get("ratingScore") else None,
            "reviews":        int(item["review"]) if item.get("review") else None,
            "sold_per_month": parse_sold(item.get("itemSoldCntShow")),
            "in_stock":       item.get("inStock", False),
            "location":       item.get("location", ""),
            "sku_id":         item.get("skuId"),
            "image":          item.get("image", ""),
            "url":            url,
            "scraped_at":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    return products


async def fetch_page(session, shop_slug, page):
    try:
        async with session.get(
            build_ajax_url(shop_slug, page),
            headers=HTTP_HEADERS,
            timeout=aiohttp.ClientTimeout(total=20)
        ) as resp:
            if resp.status != 200:
                return None, f"HTTP {resp.status}"
            return json.loads(await resp.text()), None
    except Exception as e:
        return None, str(e)


async def scrape_all(shops, max_pages, min_delay, max_delay,
                     log_cb, progress_cb, stop_event):
    all_products = []
    async with aiohttp.ClientSession() as session:
        for shop_idx, shop_slug in enumerate(shops):
            if stop_event.is_set():
                log_cb(f"Stopped before {shop_slug}"); break
            log_cb(f"Starting: {shop_slug}")
            shop_products = []
            for page in range(1, max_pages + 1):
                if stop_event.is_set():
                    log_cb(f"Stopped mid-scrape of {shop_slug}"); break
                data, err = await fetch_page(session, shop_slug, page)
                if data is None:
                    log_cb(f"[{shop_slug}] page {page} failed: {err}"); break
                raw = data.get("mods", {}).get("listItems", [])
                if not raw:
                    log_cb(f"[{shop_slug}] page {page} empty — done"); break
                products = parse_products(data, shop_slug)
                shop_products.extend(products)
                log_cb(f"[{shop_slug}] page {page} — {len(products)}/{len(raw)} with Sold/Month")
                if len(products) == 0: break
                progress_cb(shop_idx, len(shops), page, max_pages)
                await asyncio.sleep(random.uniform(min_delay, max_delay))
            all_products.extend(shop_products)
            log_cb(f"Done: {shop_slug} — {len(shop_products)} products")
    return all_products


def write_summary_sheet(ws, products, shops):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Daraz Shop Scraper — Summary Report"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = XL_CENTER; ws["A1"].fill = ACCENT_FILL
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   "
                f"Shops: {len(shops)}   |   Total Products: {len(products)}")
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="555555")
    ws["A2"].alignment = XL_CENTER; ws.row_dimensions[2].height = 18
    by_shop = defaultdict(list)
    for p in products: by_shop[p["shop"]].append(p)
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Per-Shop Overview").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=row, column=1).fill = SUBHEAD_FILL; ws.cell(row=row, column=1).alignment = XL_CENTER
    ws.row_dimensions[row].height = 22; row += 1
    for col, h in enumerate(["Shop","Products","Total Sold/Mo","Avg Price (৳)","Avg Discount","Avg Rating","Top Brand","Top Product (by Sold)"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = XL_CENTER; c.border = THIN_BORDER
    ws.row_dimensions[row].height = 20; row += 1
    for i, (shop, prods) in enumerate(by_shop.items()):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        sold_vals   = [p["sold_per_month"] for p in prods if p["sold_per_month"]]
        price_vals  = [p["price"] for p in prods if p["price"]]
        rating_vals = [p["rating"] for p in prods if p["rating"]]
        disc_vals   = []
        for p in prods:
            d = p.get("discount", "")
            if d and d not in ("", "0%"):
                try: disc_vals.append(int(d.replace("%","").replace("-","").replace("+","")))
                except: pass
        brand_counts = defaultdict(int)
        for p in prods:
            if p["brand"] and p["brand"] != "N/A": brand_counts[p["brand"]] += 1
        top_brand = max(brand_counts, key=brand_counts.get) if brand_counts else "N/A"
        top_prod  = max(prods, key=lambda x: x["sold_per_month"] or 0)
        vals = [shop, len(prods), sum(sold_vals),
                round(sum(price_vals)/len(price_vals)) if price_vals else 0,
                f"{round(sum(disc_vals)/len(disc_vals))}%" if disc_vals else "N/A",
                round(sum(rating_vals)/len(rating_vals), 1) if rating_vals else "N/A",
                top_brand,
                top_prod["name"][:60] + ("..." if len(top_prod["name"]) > 60 else "")]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.font = NORMAL_FONT; c.border = THIN_BORDER
            c.alignment = XL_LEFT if col in (1, 8) else XL_CENTER
        ws.row_dimensions[row].height = 18; row += 1
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Top 10 Products by Sold/Month (All Shops)").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=row, column=1).fill = SUBHEAD_FILL; ws.cell(row=row, column=1).alignment = XL_CENTER
    ws.row_dimensions[row].height = 22; row += 1
    for col, h in enumerate(["#","Shop","Name","Brand","Price (৳)","Discount","Rating","Sold/Month"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = XL_CENTER; c.border = THIN_BORDER
    ws.row_dimensions[row].height = 20; row += 1
    for i, p in enumerate(sorted(products, key=lambda x: x["sold_per_month"] or 0, reverse=True)[:10], 1):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for col, val in enumerate([i, p["shop"], p["name"][:60]+("..." if len(p["name"])>60 else ""),
                                    p["brand"], p["price"], p["discount"],
                                    p["rating"] or "N/A", p["sold_per_month"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.font = NORMAL_FONT; c.border = THIN_BORDER
            c.alignment = XL_LEFT if col in (3, 4) else XL_CENTER
        ws.row_dimensions[row].height = 18; row += 1
    for col, w in enumerate([6,22,60,16,14,12,10,14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def write_raw_sheet(ws, products):
    ws.title = "Raw Data"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    columns = [
        ("Shop","shop",18), ("Item ID","item_id",14), ("Name","name",55),
        ("Brand","brand",14), ("Seller","seller",22), ("Price (৳)","price",12),
        ("Orig Price (৳)","original_price",14), ("Discount","discount",10),
        ("Rating","rating",10), ("Reviews","reviews",10), ("Sold/Month","sold_per_month",13),
        ("In Stock","in_stock",10), ("Location","location",12),
        ("SKU ID","sku_id",18), ("URL","url",40), ("Scraped At","scraped_at",20),
    ]
    for col, (label, _, width) in enumerate(columns, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = XL_CENTER; c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    for row_idx, p in enumerate(products, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col, (_, key, _) in enumerate(columns, 1):
            val = p.get(key)
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = fill; c.border = THIN_BORDER
            if key == "url" and val:
                c.font = LINK_FONT; c.hyperlink = val; c.alignment = XL_LEFT
            elif key == "in_stock":
                c.value = "Yes" if val else "No"
                c.font = Font(name="Arial", size=10, color="00703C" if val else "C00000", bold=True)
                c.alignment = XL_CENTER
            elif key in ("price","original_price","sold_per_month","reviews"):
                c.font = NORMAL_FONT; c.alignment = XL_RIGHT
            elif key == "rating":
                c.font = NORMAL_FONT; c.alignment = XL_CENTER
            else:
                c.font = NORMAL_FONT; c.alignment = XL_LEFT
        ws.row_dimensions[row_idx].height = 16


def save_excel(products, filepath, shops):
    if not products:
        return False, "No products to save."
    try:
        wb = Workbook()
        write_summary_sheet(wb.active, products, shops)
        write_raw_sheet(wb.create_sheet(), products)
        wb.save(filepath)
        return True, filepath
    except Exception as e:
        return False, str(e)


# ══════════════════════════════ THEME ════════════════════════════════

C_BG       = "#13171F"
C_SURFACE  = "#1A1F2B"
C_SURFACE2 = "#1F2535"
C_BORDER   = "#2C3347"
C_ACCENT   = "#4F8EF7"
C_ACCENTDK = "#3A6FD4"
C_GREEN    = "#3DD68C"
C_GREENDK  = "#27B872"
C_RED      = "#F05050"
C_TEXT     = "#E4EAF6"
C_MUTED    = "#6B7A99"
C_MONO     = "#A8C0E8"
C_ENTRY    = "#0F1623"

F_TITLE  = ("Segoe UI Semibold", 13)
F_LABEL  = ("Segoe UI",           9)
F_LABELB = ("Segoe UI Semibold",  9)
F_INPUT  = ("Segoe UI",          10)
F_MONO   = ("Consolas",           9)
F_BTNLG  = ("Segoe UI Semibold", 11)
F_BTN    = ("Segoe UI Semibold",  9)
F_STATN  = ("Segoe UI",           8)
F_STATV  = ("Segoe UI Semibold", 15)


# ═══════════════════════════ WIDGETS ═════════════════════════════════

class FlatEntry(tk.Entry):
    def __init__(self, parent, **kw):
        super().__init__(
            parent,
            bg=C_ENTRY, fg=C_TEXT,
            insertbackground=C_ACCENT,
            selectbackground=C_ACCENTDK,
            relief="flat", bd=0,
            font=kw.pop("font", F_INPUT),
            highlightthickness=1,
            highlightbackground=C_BORDER,
            highlightcolor=C_ACCENT,
            **kw
        )


class FlatButton(tk.Button):
    def __init__(self, parent, text, command, bg, fg=C_TEXT,
                 hover=None, font=F_BTN, padx=14, pady=7, **kw):
        self._bg    = bg
        self._hover = hover or C_ACCENTDK
        super().__init__(
            parent, text=text, command=command,
            bg=bg, fg=fg, font=font,
            relief="flat", bd=0,
            padx=padx, pady=pady,
            cursor="hand2",
            activebackground=self._hover,
            activeforeground=fg,
            **kw
        )
        self.bind("<Enter>", lambda e: self._hi())
        self.bind("<Leave>", lambda e: self._lo())

    def _hi(self):
        if str(self["state"]) != "disabled":
            self.config(bg=self._hover)

    def _lo(self):
        if str(self["state"]) != "disabled":
            self.config(bg=self._bg)


class SectionLabel(tk.Frame):
    """Label + hairline rule."""
    def __init__(self, parent, text, **kw):
        super().__init__(parent, bg=C_BG, **kw)
        tk.Label(self, text=text, font=F_LABELB,
                 bg=C_BG, fg=C_MUTED).pack(side="left", padx=(0, 8))
        tk.Frame(self, bg=C_BORDER, height=1).pack(side="left", fill="x", expand=True)


class Card(tk.Frame):
    def __init__(self, parent, padx=16, pady=14, **kw):
        super().__init__(
            parent,
            bg=C_SURFACE,
            highlightbackground=C_BORDER,
            highlightthickness=1,
            padx=padx, pady=pady,
            **kw
        )


class StatBox(tk.Frame):
    def __init__(self, parent, label, var, **kw):
        super().__init__(parent, bg=C_SURFACE2, **kw)
        tk.Label(self, text=label, font=F_STATN,
                 bg=C_SURFACE2, fg=C_MUTED).pack(pady=(8, 0))
        tk.Label(self, textvariable=var, font=F_STATV,
                 bg=C_SURFACE2, fg=C_ACCENT).pack(pady=(1, 8))


class SpinRow(tk.Frame):
    """Label + Spinbox — clean, no ugly Scale widget."""
    def __init__(self, parent, label, var, from_, to, increment=1, fmt=None):
        super().__init__(parent, bg=C_SURFACE)
        tk.Label(self, text=label, font=F_LABEL,
                 bg=C_SURFACE, fg=C_MUTED, width=16, anchor="w").pack(side="left")
        self._spin = tk.Spinbox(
            self, textvariable=var,
            from_=from_, to=to, increment=increment,
            format=fmt or "%.0f",
            font=F_MONO,
            bg=C_ENTRY, fg=C_MONO,
            buttonbackground=C_SURFACE2,
            relief="flat", bd=0,
            insertbackground=C_ACCENT,
            highlightthickness=1,
            highlightbackground=C_BORDER,
            highlightcolor=C_ACCENT,
            width=7,
        )
        self._spin.pack(side="left", padx=(0, 8))
        unit = "pages" if "pages" in label.lower() else "sec"
        tk.Label(self, text=unit, font=F_LABEL,
                 bg=C_SURFACE, fg=C_MUTED).pack(side="left")


# ══════════════════════════ MAIN APP ═════════════════════════════════

class DarazScraperApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Daraz Shop Scraper")
        self.configure(bg=C_BG)
        self.geometry("980x700")
        self.minsize(840, 580)

        self._stop_event = threading.Event()
        self._running    = False
        self._products   = []
        self._shop_vars  = []

        self._ttk_styles()
        self._build_ui()
        self._center()

    def _ttk_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("D.Horizontal.TProgressbar",
                    troughcolor=C_ENTRY, background=C_ACCENT,
                    bordercolor=C_BG, lightcolor=C_ACCENT,
                    darkcolor=C_ACCENTDK, thickness=5)
        s.configure("D.Vertical.TScrollbar",
                    background=C_SURFACE2, troughcolor=C_ENTRY,
                    bordercolor=C_BG, arrowcolor=C_MUTED, relief="flat")
        s.map("D.Vertical.TScrollbar",
              background=[("active", C_BORDER)])

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    # ── root layout ─────────────────────────────────────────────────

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=C_SURFACE, height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Frame(hdr, bg=C_ACCENT, width=3).pack(side="left", fill="y")
        tk.Label(hdr, text="Daraz Shop Scraper",
                 font=F_TITLE, bg=C_SURFACE, fg=C_TEXT).pack(side="left", padx=16)
        self._status_var = tk.StringVar(value="Ready")
        tk.Label(hdr, textvariable=self._status_var,
                 font=F_LABEL, bg=C_SURFACE, fg=C_MUTED).pack(side="right", padx=16)
        tk.Frame(hdr, bg=C_BORDER, height=1).pack(side="bottom", fill="x")

        # Two-column body
        body = tk.Frame(self, bg=C_BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        body.columnconfigure(0, weight=3, minsize=300)
        body.columnconfigure(1, weight=4)
        body.rowconfigure(0, weight=1)

        left  = tk.Frame(body, bg=C_BG)
        right = tk.Frame(body, bg=C_BG)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        right.grid(row=0, column=1, sticky="nsew")

        self._build_left(left)
        self._build_right(right)

    # ── left column ─────────────────────────────────────────────────

    def _build_left(self, p):
        # Shops
        SectionLabel(p, "SHOPS TO SCRAPE").pack(fill="x", pady=(0, 6))
        shops_card = Card(p, padx=14, pady=12)
        shops_card.pack(fill="x")
        self._shops_inner = tk.Frame(shops_card, bg=C_SURFACE)
        self._shops_inner.pack(fill="x")
        for slug in ["tvhutandelectronics", "shaman-international"]:
            self._add_shop_row(slug)
        tk.Frame(shops_card, bg=C_BORDER, height=1).pack(fill="x", pady=(10, 8))
        FlatButton(shops_card, "+ Add shop", self._add_shop_row,
                   bg=C_SURFACE2, fg=C_ACCENT, hover=C_BORDER,
                   font=F_BTN, padx=10, pady=5).pack(anchor="w")

        # Output
        SectionLabel(p, "OUTPUT FILE").pack(fill="x", pady=(16, 6))
        out_card = Card(p, padx=14, pady=12)
        out_card.pack(fill="x")

        tk.Label(out_card, text="Save folder", font=F_LABEL,
                 bg=C_SURFACE, fg=C_MUTED).pack(anchor="w")
        fr = tk.Frame(out_card, bg=C_SURFACE)
        fr.pack(fill="x", pady=(3, 10))
        self._dest_var = tk.StringVar(value=os.path.expanduser("~\\Desktop"))
        FlatEntry(fr, textvariable=self._dest_var,
                  font=F_MONO).pack(side="left", fill="x", expand=True, ipady=5)
        FlatButton(fr, "Browse", self._browse_dest,
                   bg=C_SURFACE2, fg=C_MUTED, hover=C_BORDER,
                   font=F_BTN, padx=10, pady=4).pack(side="left", padx=(6, 0))

        tk.Label(out_card, text="Filename", font=F_LABEL,
                 bg=C_SURFACE, fg=C_MUTED).pack(anchor="w")
        self._fname_var = tk.StringVar(
            value=f"daraz_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        FlatEntry(out_card, textvariable=self._fname_var,
                  font=F_MONO).pack(fill="x", pady=(3, 0), ipady=5)

        # Settings
        SectionLabel(p, "SETTINGS").pack(fill="x", pady=(16, 6))
        cfg = Card(p, padx=14, pady=12)
        cfg.pack(fill="x")
        self._max_pages_var = tk.IntVar(value=50)
        self._min_delay_var = tk.DoubleVar(value=1.5)
        self._max_delay_var = tk.DoubleVar(value=3.5)
        SpinRow(cfg, "Max pages",  self._max_pages_var, 1,   100, 1).pack(fill="x", pady=4)
        SpinRow(cfg, "Min delay",  self._min_delay_var, 0.5, 10, 0.5, "%.1f").pack(fill="x", pady=4)
        SpinRow(cfg, "Max delay",  self._max_delay_var, 0.5, 10, 0.5, "%.1f").pack(fill="x", pady=4)

        # Buttons
        btns = tk.Frame(p, bg=C_BG)
        btns.pack(fill="x", pady=(18, 0))
        self._run_btn = FlatButton(
            btns, "▶   Start scraping", self._start_scrape,
            bg=C_GREEN, fg="#081A0F", hover=C_GREENDK,
            font=F_BTNLG, padx=0, pady=13)
        self._run_btn.pack(fill="x", pady=(0, 6))

        self._stop_btn = FlatButton(
            btns, "■   Stop", self._stop_scrape,
            bg=C_SURFACE2, fg=C_RED, hover=C_BORDER,
            font=F_BTNLG, padx=0, pady=10)
        self._stop_btn.pack(fill="x", pady=(0, 6))
        self._stop_btn.config(state="disabled")

        FlatButton(
            btns, "Open output folder", self._open_folder,
            bg=C_SURFACE, fg=C_MUTED, hover=C_SURFACE2,
            font=F_BTN, padx=0, pady=8
        ).pack(fill="x")

    # ── right column ────────────────────────────────────────────────

    def _build_right(self, p):
        p.rowconfigure(2, weight=1)

        # Progress
        SectionLabel(p, "PROGRESS").pack(fill="x", pady=(0, 6))
        prog_card = Card(p, padx=16, pady=14)
        prog_card.pack(fill="x")

        bar_row = tk.Frame(prog_card, bg=C_SURFACE)
        bar_row.pack(fill="x", pady=(0, 12))
        self._progress = ttk.Progressbar(
            bar_row, mode="determinate",
            style="D.Horizontal.TProgressbar")
        self._progress.pack(side="left", fill="x", expand=True, pady=1)
        self._prog_pct = tk.StringVar(value="0%")
        tk.Label(bar_row, textvariable=self._prog_pct,
                 font=F_LABELB, bg=C_SURFACE, fg=C_ACCENT,
                 width=5, anchor="e").pack(side="right", padx=(10, 0))

        stats = tk.Frame(prog_card, bg=C_SURFACE)
        stats.pack(fill="x")
        self._stat_products = tk.StringVar(value="—")
        self._stat_shop     = tk.StringVar(value="—")
        self._stat_page     = tk.StringVar(value="—")
        for col, (lbl, var) in enumerate([
            ("Products found", self._stat_products),
            ("Current shop",   self._stat_shop),
            ("Page",           self._stat_page),
        ]):
            stats.columnconfigure(col, weight=1)
            StatBox(stats, lbl, var).grid(
                row=0, column=col, sticky="ew",
                padx=(0, 6) if col < 2 else 0)

        # Log
        log_hdr = tk.Frame(p, bg=C_BG)
        log_hdr.pack(fill="x", pady=(16, 6))
        SectionLabel(log_hdr, "ACTIVITY LOG").pack(side="left", fill="x", expand=True)
        FlatButton(log_hdr, "Clear", self._clear_log,
                   bg=C_SURFACE, fg=C_MUTED, hover=C_SURFACE2,
                   font=F_BTN, padx=10, pady=3).pack(side="right")

        log_wrap = tk.Frame(p, bg=C_SURFACE,
                            highlightbackground=C_BORDER, highlightthickness=1)
        log_wrap.pack(fill="both", expand=True)
        log_wrap.rowconfigure(0, weight=1)
        log_wrap.columnconfigure(0, weight=1)

        self._log = tk.Text(
            log_wrap,
            bg=C_ENTRY, fg=C_TEXT,
            font=F_MONO,
            relief="flat", bd=0,
            padx=12, pady=10,
            insertbackground=C_ACCENT,
            selectbackground=C_ACCENTDK,
            state="disabled", wrap="word",
            spacing1=1, spacing3=2,
        )
        vsb = ttk.Scrollbar(log_wrap, orient="vertical",
                            command=self._log.yview,
                            style="D.Vertical.TScrollbar")
        self._log.configure(yscrollcommand=vsb.set)
        self._log.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self._log.tag_config("info",    foreground=C_TEXT)
        self._log.tag_config("success", foreground=C_GREEN)
        self._log.tag_config("warn",    foreground="#E8B84B")
        self._log.tag_config("error",   foreground=C_RED)
        self._log.tag_config("accent",  foreground=C_ACCENT)
        self._log.tag_config("ts",      foreground=C_MUTED)

    # ── shop rows ───────────────────────────────────────────────────

    def _add_shop_row(self, slug=""):
        row = tk.Frame(self._shops_inner, bg=C_SURFACE)
        row.pack(fill="x", pady=2)

        enabled = tk.BooleanVar(value=True)

        pill = tk.Label(row, text="ON", font=("Segoe UI Semibold", 7),
                        bg=C_GREEN, fg="#081A0F",
                        padx=7, pady=2, cursor="hand2")
        pill.pack(side="left", padx=(0, 8))

        def toggle():
            if enabled.get():
                enabled.set(False)
                pill.config(text="OFF", bg=C_SURFACE2, fg=C_MUTED)
            else:
                enabled.set(True)
                pill.config(text="ON", bg=C_GREEN, fg="#081A0F")

        pill.bind("<Button-1>", lambda e: toggle())

        slug_var = tk.StringVar(value=slug)
        FlatEntry(row, textvariable=slug_var,
                  font=F_MONO).pack(side="left", fill="x", expand=True, ipady=4)

        def remove():
            self._shop_vars = [(s, e) for s, e in self._shop_vars if s is not slug_var]
            row.destroy()

        FlatButton(row, "✕", remove,
                   bg=C_SURFACE, fg=C_MUTED, hover=C_SURFACE2,
                   font=("Segoe UI", 8), padx=8, pady=4
                   ).pack(side="left", padx=(6, 0))

        self._shop_vars.append((slug_var, enabled))

    # ── helpers ─────────────────────────────────────────────────────

    def _log_msg(self, msg, tag="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.config(state="normal")
        self._log.insert("end", f" {ts}  ", "ts")
        self._log.insert("end", f"{msg}\n", tag)
        self._log.see("end")
        self._log.config(state="disabled")

    def _clear_log(self):
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")

    def _browse_dest(self):
        folder = filedialog.askdirectory(
            initialdir=self._dest_var.get(),
            title="Choose output folder")
        if folder:
            self._dest_var.set(folder)

    def _open_folder(self):
        path = self._dest_var.get()
        if os.path.isdir(path):
            os.startfile(path)
        else:
            messagebox.showwarning("Not found", f"Folder not found:\n{path}")

    def _set_running(self, running):
        self._running = running
        self._run_btn.config(state="disabled" if running else "normal")
        self._stop_btn.config(
            state="normal" if running else "disabled",
            bg=C_RED if running else C_SURFACE2,
            fg=C_TEXT if running else C_RED)

    # ── scraping ────────────────────────────────────────────────────

    def _start_scrape(self):
        shops = [s.get().strip() for s, e in self._shop_vars
                 if e.get() and s.get().strip()]
        if not shops:
            messagebox.showwarning("No shops",
                "Add at least one shop slug and make sure it's enabled.")
            return

        dest  = self._dest_var.get().strip()
        fname = self._fname_var.get().strip()
        if not fname.endswith(".xlsx"): fname += ".xlsx"
        output_path = os.path.join(dest, fname)

        max_pages = self._max_pages_var.get()
        min_delay = self._min_delay_var.get()
        max_delay = self._max_delay_var.get()
        if min_delay > max_delay:
            min_delay, max_delay = max_delay, min_delay

        self._stop_event.clear()
        self._set_running(True)
        self._products = []
        self._progress["value"] = 0
        self._prog_pct.set("0%")
        self._stat_products.set("0")
        self._stat_shop.set("—")
        self._stat_page.set("—")
        self._status_var.set(f"Scraping {len(shops)} shop(s)...")
        self._log_msg(f"Starting — {len(shops)} shop(s)", "accent")
        for s in shops:
            self._log_msg(f"  · {s}")

        def on_log(msg):
            tag = ("success" if any(x in msg for x in ("Done:", "empty")) else
                   "warn"    if "failed" in msg or "Stopped" in msg else
                   "info")
            self.after(0, lambda m=msg, t=tag: self._log_msg(m, t))

        def on_progress(shop_idx, total_shops, page, max_p):
            pct = ((shop_idx + page / max_p) / total_shops) * 100
            self.after(0, lambda: [
                self._progress.__setitem__("value", pct),
                self._prog_pct.set(f"{int(pct)}%"),
                self._stat_shop.set(shops[shop_idx]),
                self._stat_page.set(f"{page}/{max_p}"),
                self._stat_products.set(str(len(self._products))),
            ])

        def worker():
            try:
                products = asyncio.run(scrape_all(
                    shops, max_pages, min_delay, max_delay,
                    on_log, on_progress, self._stop_event))
                self._products = products
                self.after(0, lambda: self._stat_products.set(str(len(products))))

                if not self._stop_event.is_set():
                    ok, result = save_excel(products, output_path, shops)
                    if ok:
                        self.after(0, lambda: [
                            self._log_msg(
                                f"Saved {len(products)} products to {output_path}",
                                "success"),
                            self._progress.__setitem__("value", 100),
                            self._prog_pct.set("100%"),
                            self._status_var.set(f"Done — {len(products)} products"),
                            messagebox.showinfo("Done",
                                f"Scraped {len(products)} products.\n\nSaved to:\n{output_path}"),
                        ])
                    else:
                        self.after(0, lambda r=result: [
                            self._log_msg(f"Save failed: {r}", "error"),
                            self._status_var.set("Save failed"),
                        ])
                else:
                    self.after(0, lambda: self._status_var.set("Stopped"))
            except Exception as ex:
                self.after(0, lambda e=ex: [
                    self._log_msg(f"Fatal error: {e}", "error"),
                    self._status_var.set("Error"),
                ])
            finally:
                self.after(0, lambda: self._set_running(False))

        threading.Thread(target=worker, daemon=True).start()

    def _stop_scrape(self):
        self._stop_event.set()
        self._log_msg("Stop requested...", "warn")
        self._status_var.set("Stopping...")


# ══════════════════════════ ENTRY POINT ══════════════════════════════

def main():
    app = DarazScraperApp()
    app.mainloop()


if __name__ == "__main__":
    main()
